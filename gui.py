import tkinter as tk
from tkinter import filedialog, messagebox
import os
import re
import openpyxl
from collections import Counter

is_excel_uploaded = False
is_all_parameters_set = False
file_path = None
size_deger_list = [] # [size , res value , PN , power rating]
pwr_list = []

r_size = '0402'
r_pwr = 50
bias_pwr = 100
prl_cnt = 3
seri_cnt = 3
vref = 2.5
vout = 120
v_tol = 0.001
derating_rate = 0.8

def are_lists_equal(list1, list2):
    new_list1 = []
    new_list2 = []
    for i in list1:
        new_list1.append(i[1])
    for i in list2:
        new_list2.append(i[1])
    return Counter(new_list1) == Counter(new_list2)

def comp_out(vref, rtop, rbot):
    # global vout
    if rbot == 0:
        return -1
    vout = vref*(1+rtop/rbot)
    return vout

def comp_prl(prl_lst):
    lgt = len(prl_lst)
    req = prl_lst[0][0]
    for i in range(lgt - 1):
        req = prl_lst[i+1][0]*req/(req + prl_lst[i+1][0])
    return req
        
def comp_seri(seri_list):
    val = 0
    for i in seri_list:
        val += i[0]
    return val

def parallelfinder(deger_list , ctr, mem):
    global size_deger_list
    if ctr != 1:
        for i in range(len(size_deger_list)):
            deger_list.append([size_deger_list[i][1],size_deger_list[i][2], size_deger_list[i][3]])
            parallelfinder(deger_list, ctr - 1, mem)
            deger_list.pop()
    else:
        for i in range(len(size_deger_list)):
            deger_list.append([size_deger_list[i][1],size_deger_list[i][2], size_deger_list[i][3]])
            ch = True
            for j in deger_list:
                if j[0] == 0:
                    ch = False
                    break
            if ch:
                mem.append(deger_list.copy())
            deger_list.pop()
                
                

def serifinder(deger_list, ctr, mem):
    global size_deger_list
    if ctr != 1:
        for i in range(len(size_deger_list)):
            deger_list.append([size_deger_list[i][1],size_deger_list[i][2], size_deger_list[i][3]])
            serifinder(deger_list, ctr - 1, mem)
            deger_list.pop()
    else:
        for i in range(len(size_deger_list)):
            deger_list.append([size_deger_list[i][1],size_deger_list[i][2], size_deger_list[i][3]])
            mem.append(deger_list.copy())
            # print("size: ", len(deger_list))
            deger_list.pop()
            
def compPwr(vol, res): 
    return vol*vol/res
          
def check_pwr_ratings(tmp_vout, seri_val, prl_val, r_pwr, bias_pwr):
    req_prl = comp_prl(prl_val)
    req_ser = comp_seri(seri_val)
    prl_vol = tmp_vout * req_prl / (req_prl + req_ser)
    seri_vol_total = tmp_vout * req_ser / (req_prl + req_ser)
    if prl_vol >= r_pwr:
        return False
    prl_pwrs = []
    for i in prl_val:
        tmp_pwr = compPwr(prl_vol, i[0])
        if tmp_pwr >= derating_rate * i[2]:
            return False
    seri_current = tmp_vout / (req_ser + req_prl)
    for i in seri_val:
        tmp_pwr = i[0] * seri_current * seri_current
        if tmp_pwr >= derating_rate * i[2]:
            return False
        tmp_vol = seri_vol_total * i[0] / req_ser
        if tmp_vol >= r_pwr:
            return False
    total_pwr = seri_current*seri_current*(req_prl + req_ser)
    if total_pwr > bias_pwr/1000:
        return False
    return True

def calc_bias_pwr(tmp_vout, seri_val, prl_val, r_pwr, bias_pwr):
    req_prl = comp_prl(prl_val)
    req_ser = comp_seri(seri_val)
    prl_vol = tmp_vout * req_prl / (req_prl + req_ser)
    seri_vol_total = tmp_vout * req_ser / (req_prl + req_ser)
    seri_current = tmp_vout / (req_ser + req_prl)
    total_pwr = seri_current*seri_current*(req_prl + req_ser)
    return total_pwr
    
# Decorator to check if Excel is uploaded and validate inputs
def check_excel_and_validate(func):
    def wrapper(*args, **kwargs):
        global is_excel_uploaded
        if not is_excel_uploaded:
            messagebox.showerror("Error", "You haven't uploaded an Excel file!")
        else:
            try:
                # Perform input validation
                get_values()
                func(*args, **kwargs)
                # print("func is over!!!")
            except ValueError as e:
                messagebox.showerror("Error", str(e))  # Display validation error in a message box

    return wrapper

# Function to set the initial directory for the file dialog
def get_initial_directory():
    script_directory = os.path.dirname(os.path.realpath(__file__))
    return script_directory

def upload_excel():
    global is_excel_uploaded
    global file_path
    initial_dir = get_initial_directory()
    file_path = filedialog.askopenfilename(initialdir=initial_dir)
    if file_path:  # Check if a file is selected
        is_excel_uploaded = True
        # Do something with the file_path (like reading the Excel file)
    # Do something with the file_path (like reading the Excel file)

#tmp_list tol, vout, bias_pwr, seri_codes, prl_codes
def check_res_list(tmp_list, res_list_tol, res_list_pwr):
    
    if(len(res_list_tol) == 0):
        res_list_tol.append(tmp_list)
    else:
        ch = True
        for i in res_list_tol:
            # print(i[3])
            # print(tmp_list[3])
            if are_lists_equal(i[3], tmp_list[3]) and are_lists_equal(i[4], tmp_list[4]):
                ch = False
                break
        if ch:
            for i in range(len(res_list_tol)):
                if res_list_tol[i][0] > tmp_list[0]:
                    res_list_tol.insert(i, tmp_list.copy())
                    break
    if(len(res_list_pwr) == 0):
        res_list_pwr.append(tmp_list)
    else:
        ch = True
        for i in res_list_pwr:
            if are_lists_equal(i[3], tmp_list[3]) and are_lists_equal(i[4], tmp_list[4]):
                ch = False
                break
        if ch:
            for i in range(len(res_list_pwr)):
                if res_list_pwr[i][2] > tmp_list[2]:
                    res_list_pwr.insert(i, tmp_list.copy())
                    break
                    # print("res_list_pwr: ")
                    # print(res_list_pwr)
    if len(res_list_tol) > 5:
        res_list_tol.pop()
    if len(res_list_pwr) > 5:
        res_list_pwr.pop()
    return res_list_tol, res_list_pwr
    
def display_results(res_list_tol, res_list_pwr):
    global prl_cnt
    global seri_cnt
    row_space = max(prl_cnt, seri_cnt)
    result_window = tk.Toplevel(root)
    res_list_pwr_len = len(res_list_pwr)
    result_window.title("Results")
    res_list_tol_len = len(res_list_tol)
    canvas = tk.Canvas(result_window, width=360, height= 400 + int((res_list_tol_len + res_list_pwr_len) * row_space * 20))
    canvas.pack()
    # print("res_list_tol: ")
    # print(res_list_tol)
    # Display res_list_tol
    canvas.create_text(200, 20, text="Results for accuracy", font=("Helvetica", 14), anchor="n")
    for idx, result in enumerate(res_list_tol, start=1):
        display_text = f"{idx}. Vout: {result[1]:.2f}, Tolerance: {result[0]:.4f}, Bias Power: {result[2]:.2f}"
        canvas.create_text(10, 40 + (idx * 20), text=display_text, anchor="w")

    # Display res_list_pwr
    canvas.create_text(200, 160, text="Results for efficiency", font=("Helvetica", 14), anchor="n")
    for idx, result in enumerate(res_list_pwr, start=1):
        display_text = f"{idx}. Vout: {result[1]:.2f}, Tolerance: {result[0]:.4f}, Bias Power: {result[2]:.2f}"
        canvas.create_text(10, 180 + (idx * 20), text=display_text, anchor="w")

    canvas.create_text(200, 300, text="Serial Codes       ||    Parallel Codes", font=("Helvetica", 14), anchor="n")
    space = 0
    for idx, result in enumerate(res_list_tol, start=1):
        display_text = f"{idx}"
        canvas.create_text(10, 360 + space, text=display_text, anchor="w")
        ser_space = 0
        for i_idx in range(len(result[3])):
            display_text = f"Code: {result[3][i_idx]}"
            canvas.create_text(20, 360 + space + (i_idx * 20), text=display_text, anchor="w")
            ser_space += 20
        par_space = 0
        for i_idx in range(len(result[4])):
            display_text = f"Code: {result[4][i_idx]}"
            canvas.create_text(210, 360 + space + (i_idx * 20), text=display_text, anchor="w")
            par_space += 20
        space += max(ser_space, par_space)
    stt = int(360 + res_list_tol_len * row_space * 20 + 40)
    space = 0
    for idx, result in enumerate(res_list_pwr, start=1):
        display_text = f"{idx}"
        canvas.create_text(10, stt + space, text=display_text, anchor="w")
        ser_space = 0
        for i_idx in range(len(result[3])):
            display_text = f"Code: {result[3][i_idx]}"
            canvas.create_text(20, stt + space + (i_idx * 20), text=display_text, anchor="w")
            ser_space += 20
        par_space = 0
        for i_idx in range(len(result[4])):
            display_text = f"Code: {result[4][i_idx]}"
            canvas.create_text(210, stt + space + (i_idx * 20), text=display_text, anchor="w")
            par_space += 20
        space += max(ser_space, par_space)
    
@check_excel_and_validate
def calculate():
    global size_deger_list
    global pwr_list
    global v_tol
    global vout
    global vref
    global seri_cnt
    global prl_cnt
    try:
        res_list = []
        if not is_excel_uploaded:
            print("exceli yüklemedin!")
        if not is_all_parameters_set:
            print("bütün parametreleri düzgün gir!")
        excel_wb = openpyxl.load_workbook(file_path)
        çalışma_sayfası = excel_wb.active  # Aktif çalışma sayfasını kullanabilirsiniz
        veri_listesi = []
        # Perform calculations based on the inputs
        # Access the input values using the corresponding variables
        # Belirli bir kolondaki tüm hücreleri döngü kullanarak alın
        for satır_index, satır in enumerate(çalışma_sayfası.iter_rows(values_only=True, min_col=2, max_col=2), start=3):
            for hücre in satır:
                if hücre != None:
                    veri_listesi.append(hücre)

        kod_listesi = []
        for satır_index, satır in enumerate(çalışma_sayfası.iter_rows(values_only=True, min_col=1, max_col=1), start=3):
            for hücre in satır:
                if hücre != None:
                    kod_listesi.append(hücre)

        excel_wb.close()
        size_deger_list = [] # [size , res value , PN , power rating]
        pwr_list = []
        ctr = 0
        for eşleşme in veri_listesi:
            size = None
            klas = None
            if ' 0603' in eşleşme or '/0603' in eşleşme:
                size = '0603'
            if ' 0402' in eşleşme or '/0402' in eşleşme:
                size = '0402'
            if ' 0805' in eşleşme or '/0805' in eşleşme:
                size = '0805'
            if ' F ' in eşleşme or ' %1 ' in eşleşme:
                klas = 'F'
            birim = None
            if 'R ' in eşleşme:
                carpan = 1
                birim = 'R '
            if 'K ' in eşleşme:
                carpan = 1000
                birim = 'K '
            if 'M ' in eşleşme:
                carpan = 1000000
                birim = 'M '
            if size != r_size or birim == None or klas == None:
                ctr += 1
                continue
            blc = eşleşme.split(birim)
            blc_x = blc[0].split(' ')
            blc_y = blc[1].split(' ')
            chp = True
            for g in blc_y:
                if 'W' in g:
                    tmp_lst = g.split('W')
                    tmp_str = tmp_lst[0]
                    if '/' in tmp_str:
                        tmp_str = tmp_str.split('/')
                        try:
                            pwr_list.append(float(tmp_str[0])/float(tmp_str[1]))
                            chp = False
                        except:
                            chp = True
                            break
                    else:
                        try:
                            pwr_list.append(float(tmp_str))
                            chp = False
                        except:
                            chp = True
                            break
                    break
            if chp:
                continue
            deger = blc_x[-1]
            if '*' not in deger:
                deger = float(deger)
                deger = carpan * deger
                size_deger_list.append([size, deger, kod_listesi[ctr], pwr_list[-1]])
            ctr += 1
        prl_ctr = 0
        seri_mem = [] # [res value , PN , power rating]
        seri_deger_list = []
        serifinder(seri_deger_list, seri_cnt, seri_mem)
        # print("seri mem")
        # for i in seri_mem:
            # print(i)
        prl_mem = [] # [res value , PN , power rating]
        prl_deger_list = []
        parallelfinder(prl_deger_list, prl_cnt, prl_mem)
        # print("paralel mem")
        # for i in prl_mem:
            # print(i)
        res_list_tol = []
        res_list_pwr = []
        for i in seri_mem:
            seri_val = comp_seri(i)
            # print(seri_val)
            for j in prl_mem:
                prl_val = comp_prl(j)
                # print(prl_val)
                tmp_vout = comp_out(vref, seri_val, prl_val)
                if tmp_vout == -1:
                    continue
                tol = abs((tmp_vout - vout) / vout)
                # print("tmp_vout, vout: ", tmp_vout, vout)
                # print("tol, v_tol: ", tol, v_tol)
                if tol <= v_tol:
                    if check_pwr_ratings(tmp_vout, i, j, r_pwr, bias_pwr):
                        tmp_bias_pwr = calc_bias_pwr(tmp_vout, i, j, r_pwr, bias_pwr)
                        #tmp_list tol, vout, bias_pwr, seri_codes, prl_codes
                        tmp_list = [tol, tmp_vout, tmp_bias_pwr, i, j]
                        res_list_tol, res_list_pwr = check_res_list(tmp_list, res_list_tol, res_list_pwr)
                        print("vout: ", tmp_vout, " seri codes: ",  i, " paralel codes: ", j)
        # print("this is res_list_tol: ")
        # for i in res_list_tol:
            # print(i)
        # print("this is res_list_pwr: ")
        # for i in res_list_pwr:
            # print(i)
        display_results(res_list_tol, res_list_pwr)
    except Exception as e:
        print("Error in calculate function:", e)
    finally:
        print("func is over!!!")

# Create the main window
root = tk.Tk()
root.title("Resistor Calculation")

# Function to get the values from the input fields
def get_values():
    global seri_cnt
    global prl_cnt
    global r_size
    global derating_rate
    global r_pwr
    global vref
    global vout
    global bias_pwr
    try:
        # print("now running: get_values")
        parallel_resistors = int(parallel_resistors_entry.get())
        prl_cnt = parallel_resistors
        # print("prl_cnt: ", prl_cnt)
        serial_resistors = int(serial_resistors_entry.get())
        seri_cnt = serial_resistors
        # print("seri count: ", seri_cnt)
        resistor_size = resistor_size_var.get()
        r_size = resistor_size
        # print("r_size: ", r_size)
        derating = 100.0/(float(derating_entry.get())+100.0)
        derating_rate = derating
        # print("derating_rate: ", derating_rate)
        max_voltage = float(max_voltage_entry.get())
        r_pwr = max_voltage
        # print("r_pwr: ", r_pwr)
        reference_voltage = float(reference_voltage_entry.get())
        vref = reference_voltage
        # print("vref: ", vref)
        output_voltage = float(output_voltage_entry.get())
        vout = output_voltage
        # print("vout: ", vout)
        max_power = float(max_power_entry.get())
        # print("max_power: ", max_power)
        bias_pwr = max_power

        # Check if any field is empty or not in the right format
        if (parallel_resistors <= 0 or
                serial_resistors <= 0 or
                resistor_size not in ['0402', '0603', '0805'] or
                max_voltage <= 0 or
                derating <= 0 or
                derating > 1 or
                reference_voltage <= 0 or
                output_voltage <= 0 or
                max_power <= 0):
            raise ValueError("Invalid input. Please ensure all values are filled correctly.")
            messagebox.showerror("Invalid input. Please ensure all values are filled correctly.")  # Display validation error in a message box

        # All inputs are valid, you can proceed with the calculations or actions here
        # ...
        print("get_values over!!!")
    except ValueError as e:
        # Handle the validation error
        # You might display an error message, log the issue, or perform any action necessary
        print("Validation Error:", e)
        messagebox.showerror("Error", str(e))  # Display validation error in a message box

# Create and place labels and entry boxes for user input
parallel_resistors_label = tk.Label(root, text="How many parallel resistors at bottom?")
parallel_resistors_label.pack()
parallel_resistors_entry = tk.Entry(root)
parallel_resistors_entry.pack()

serial_resistors_label = tk.Label(root, text="How many serial resistors on top?")
serial_resistors_label.pack()
serial_resistors_entry = tk.Entry(root)
serial_resistors_entry.pack()

resistor_sizes = ['0402', '0603', '0805']
resistor_size_var = tk.StringVar()
resistor_size_label = tk.Label(root, text="Size of the resistors?")
resistor_size_label.pack()
resistor_size_dropdown = tk.OptionMenu(root, resistor_size_var, *resistor_sizes)
resistor_size_dropdown.pack()

max_voltage_label = tk.Label(root, text="Maximum voltage for a resistor")
max_voltage_label.pack()
max_voltage_entry = tk.Entry(root)
max_voltage_entry.pack()

derating_label = tk.Label(root, text="Derating ratio? (%)")
derating_label.pack()
derating_entry = tk.Entry(root)
derating_entry.pack()

reference_voltage_label = tk.Label(root, text="Reference Voltage")
reference_voltage_label.pack()
reference_voltage_entry = tk.Entry(root)
reference_voltage_entry.pack()

output_voltage_label = tk.Label(root, text="Output Voltage")
output_voltage_label.pack()
output_voltage_entry = tk.Entry(root)
output_voltage_entry.pack()

max_power_label = tk.Label(root, text="Maximum power of bias? (in mW)")
max_power_label.pack()
max_power_entry = tk.Entry(root)
max_power_entry.pack()

# Button to trigger the calculation
calculate_button = tk.Button(root, text="Calculate", command=calculate)
calculate_button.pack()

# Button to upload Excel
upload_button = tk.Button(root, text="Upload Excel", command=upload_excel)
upload_button.pack()

root.mainloop()